# @author Cesar Ramirez
# Optimization model for choosing class schedule
import gurobipy
from gurobipy import *

# READ DATA FROM EXCEL
import openpyxl as opxl

# Load parameters from Excel file
path = "C:\\dev\\pipe\\CourseConnect\\Schedule-Data.xlsx"
doc = opxl.load_workbook(path)

# ---- Initializations -------------

# Set of classes C
C = []

# Set of semesters
# Assumed, could be updated later with UI
S = (0, 1, 2, 3, 4, 5, 6, 7)

# Parameters

# Credit hours indexed by class c
hrs = {}

# Constants
MIN_HOURS = 12
MAX_HOURS = 20

# Prerequisites list indexed by class c
prereq = {}

# Availability list indexed by class c
avai = {}

# ----------- Read from excel ----------------

ws = doc["ISE Prereqs"]

# Populating classes set C
row = 1
while ws.cell(row=row+1, column=1).value:
    C.append(ws.cell(row=row+1, column=1).value)
    row += 1

# Populating parameter prerequisites prereq
# Indexed by class c first, then prerequisites index i from class order C
row = 1
col = 1
ind_pre = []
while ws.cell(row=row+1, column=col).value:
    ind_pre = []
    while ws.cell(row=row, column=col+1).value or ws.cell(row=row, column=col+1).value == 0:
        ind_pre.append(ws.cell(row=row+1, column=col+1).value)
        col += 1
    prereq[C[row - 1]] = ind_pre
    col = 1
    row += 1

pre = {}
for p in prereq.keys():
    index = 0
    pre.setdefault(p, [])
    for i in prereq[p]:
        if i == 1:
            pre.get(p, []).append(C[index])
        index += 1


# Populating parameter hours h
ws = doc["ISE"]
row = 1
while ws.cell(row=row+1, column=1).value:
    hrs[C[row-1]] = ws.cell(row=row+1, column=5).value
    row += 1

# Populating parameter avai
ws = doc["ISE Availability"]

row = 1
col = 1
avai = {}
ind_avai = [0, 0, 0, 0, 0, 0, 0, 0]
while ws.cell(row=row+1, column=col).value:
    ind_avai = [0, 0, 0, 0, 0, 0, 0, 0]
    for i in range(8):
        ind_avai[i] = ws.cell(row=row+1, column=col+1).value
        col += 1
    avai[C[row-1]] = ind_avai
    col = 1
    row += 1

# ---------- OPTIMIZATION MODEL ----------------------
# Model
model = Model("Schedule")
model.setParam(GRB.Param.OutputFlag, 0)


# Decision Variables

# sched - first index will be the semester (int) and the second term will be class name (string)
# 1 for taking class at class c at semester s, 0 otherwise
schedule = model.addVars(S, C, vtype=GRB.BINARY, name="schedule")

# bound - single integer variable that represents the highest number of hours that will be required in a semester
# Implicitly accounts for the min and max number of hours per semester to be a full time student
# bounds = model.addVars(S, vtype=GRB.INTEGER, lb=MIN_HOURS, ub=MAX_HOURS, name="bounds")
bound = model.addVar(vtype=GRB.INTEGER, lb=MIN_HOURS, ub=MAX_HOURS, name="bound")

bounds = {}
expr = LinExpr()
semester_count = 0
index_count = 0
for v in schedule.items():
    if v[0][0] != index_count:
        bounds.setdefault(index_count, expr)
        index_count += 1
        expr = LinExpr()
    expr.add(schedule[(v[0][0], v[0][1])], hrs[v[0][1]])

# Constraints

# Configuring bound constraint
model.addConstrs(quicksum(schedule[(s, c)]*hrs[c] for c in C) <= bound for s in S)

# Take each required class
model.addConstrs(quicksum(schedule[(s, c)] for s in S) == 1 for c in C)

# Availability constraint
model.addConstrs(schedule[(i, j)] <= avai[j][i] for i in S for j in C)

# Prerequisite constraint
for c in C:
    for p in pre.get(c, []):
        model.addConstrs(schedule[(s, p)] <= schedule[(s, c)] for s in S)

# above will work, but here is what I need:
# prereq needs to take in class (string) and return a list of prerequisite classes (strings)

# # Prerequisite constraint
# model.addConstrs(quicksum(schedule[(k, C[p])] for k in range(s)) <= prereq[c][p] * schedule[(s, c)]
#                  for p in range(len(C)) for c in C for s in S)

# # Another attempt at the prerequisite constraint...didn't work >:(
P = list(C)
# model.addConstrs(quicksum(schedule[(k, p)] for k in range(s)) >= prereq[c][int(P.index(p))] * schedule[(s, c)]
#                  for p in P for c in C for s in S)


# Objective function
model.setObjective(max(x for x in bounds.items()), sense=GRB.MINIMIZE)
model.setParam("OutputFlag", 0)

model.update()
model.optimize()

hour_count = 0

# Printing outputs
if model.status == GRB.OPTIMAL:
    print("\nMaximum number of hours required:", model.objVal)
    print("----Schedule---")
    semester = 1
    sem = semester
    sem_hours = 0
    print(f"\nSemester {semester}")
    for v in model.getVars():
        if v.x == 1:
            course = v.VarName[11:20].split("]")[0]
            sem = str(int(v.VarName[9]) + 1)
            print(f"{course}: {hrs[course]}")
            sem_hours += hrs[course]
        if int(sem) > int(semester):
            semester = sem
            print(f"-----Total Hours: {sem_hours}")
            print(f"\nSemester {semester}")
            sem_hours = 0
    # Quick fix for outputting last semester hours
    print(f"-----Total Hours: {sem_hours}")

else:
    print("Not possible :(")
