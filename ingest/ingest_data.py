# READ DATA FROM EXCEL
import openpyxl as opxl

# Load parameters from Excel file
path = "C:\\dev\\pipe\\Shaman\\Schedule-Data.xlsx"
doc = opxl.load_workbook(path)

# ---- Initializations -------------

# Set of classes C
C = []

# Set of semesters
# Assumed, could be updated later with UI
S = (0, 1, 2, 3, 4, 5, 6, 7)

# Parameters

# Credit hours indexed by class c
hrs = {}

# Constants
MIN_HOURS = 12
MAX_HOURS = 20

# Prerequisites list indexed by class c
prereq = {}

# Availability list indexed by class c
avai = {}

# ----------- Read from excel ----------------

ws = doc["Sheet1"]

# Populating classes set C
row = 1
while ws.cell(row=row+1, column=1).value:
    C.append(ws.cell(row=row+1, column=1).value)
    row += 1

# Populating parameter prerequisites prereq
# Indexed by class c first, then prerequisites index i from class order C
row = 1
col = 1
ind_pre = []
while ws.cell(row=row+1, column=col).value:
    ind_pre = []
    while ws.cell(row=row, column=col+1).value or ws.cell(row=row, column=col+1).value == 0:
        ind_pre.append(ws.cell(row=row+1, column=col+1).value)
        col += 1
    prereq[C[row - 1]] = ind_pre
    col = 1
    row += 1

pre = {}
for p in prereq.keys():
    index = 0
    pre.setdefault(p, [])
    for i in prereq[p]:
        if i == 1:
            pre.get(p, []).append(C[index])
        index += 1

# Populating parameter hours h
ws = doc["ISE"]
row = 1
while ws.cell(row=row+1, column=1).value:
    hrs[C[row-1]] = ws.cell(row=row+1, column=5).value
    row += 1

# Populating parameter avai
ws = doc["ISE Availability"]

row = 1
col = 1
avai = {}
ind_avai = [0, 0, 0, 0, 0, 0, 0, 0]
while ws.cell(row=row+1, column=col).value:
    ind_avai = [0, 0, 0, 0, 0, 0, 0, 0]
    for i in range(8):
        ind_avai[i] = ws.cell(row=row+1, column=col+1).value
        col += 1
    avai[C[row-1]] = ind_avai
    col = 1
    row += 1
