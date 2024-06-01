# @author Cesar Ramirez
# Optimization model for choosing class schedule
import gurobipy
from gurobipy import *

# READ DATA FROM EXCEL
import openpyxl as opxl

# Load parameters from Excel file
path = "C:\\dev\\pipe\\Shaman\\Schedule-Data.xlsx"
doc = opxl.load_workbook(path)

# ---- Initializations -------------

# Set of classes C
C = []

# Set of semesters
# Assumed, could be updated later with UI
S = (0, 1, 2, 3, 4, 5, 6, 7)

# Parameters

# Credit hours indexed by class c
hrs = {}

# Constants
MIN_HOURS = 12
MAX_HOURS = 20

# Prerequisites list indexed by class c
prereq = {}

# Availability list indexed by class c
avai = {}

# ----------- Read from excel ----------------

ws = doc["Sheet1"]

# Populating classes set C
row = 1
while ws.cell(row=row+1, column=1).value:
    C.append(ws.cell(row=row+1, column=1).value)
    row += 1

# Populating parameter prerequisites prereq
# Indexed by class c first, then prerequisites index i from class order C
row = 1
col = 1
ind_pre = []
while ws.cell(row=row+1, column=col).value:
    ind_pre = []
    while ws.cell(row=row, column=col+1).value or ws.cell(row=row, column=col+1).value == 0:
        ind_pre.append(ws.cell(row=row+1, column=col+1).value)
        col += 1
    prereq[C[row - 1]] = ind_pre
    col = 1
    row += 1

pre = {}
for p in prereq.keys():
    index = 0
    pre.setdefault(p, [])
    for i in prereq[p]:
        if i == 1:
            pre.get(p, []).append(C[index])
        index += 1

# Populating parameter hours h
ws = doc["ISE"]
row = 1
while ws.cell(row=row+1, column=1).value:
    hrs[C[row-1]] = ws.cell(row=row+1, column=5).value
    row += 1

# Populating parameter avai
ws = doc["ISE Availability"]

row = 1
col = 1
avai = {}
ind_avai = [0, 0, 0, 0, 0, 0, 0, 0]
while ws.cell(row=row+1, column=col).value:
    ind_avai = [0, 0, 0, 0, 0, 0, 0, 0]
    for i in range(8):
        ind_avai[i] = ws.cell(row=row+1, column=col+1).value
        col += 1
    avai[C[row-1]] = ind_avai
    col = 1
    row += 1

# ---------- OPTIMIZATION MODEL ----------------------
# Model
model = Model("Schedule")

# Decision Variables

# sched - first index will be the semester (int) and the second term will be class name (string)
# 1 for taking class at class c at semester s, 0 otherwise
schedule = model.addVars(S, C, vtype=GRB.BINARY, name="schedule")

# bound - single integer variable that represents the highest number of hours that will be required in a semester
# Implicitly accounts for the min and max number of hours per semester to be a full time student
# bounds = model.addVars(S, vtype=GRB.INTEGER, lb=MIN_HOURS, ub=MAX_HOURS, name="bounds")
bound = model.addVar(vtype=GRB.INTEGER, lb=MIN_HOURS,
                     ub=MAX_HOURS, name="bound")

# Constraints

# Configuring bound constraint
model.addConstrs(quicksum(schedule[(s, c)]*hrs[c]
                 for c in C) <= bound for s in S)

# Take each required class
model.addConstrs(quicksum(schedule[(s, c)] for s in S) == 1 for c in C)

# Availability constraint
model.addConstrs(schedule[(i, j)] <= avai[j][i] for i in S for j in C)

# Prerequisite constraint
# Note - pre[c].count(j) is 1 if class j is a prerequisite for class c
# After first semester
model.addConstrs(schedule[(s, c)] * pre[c].count(j) <= quicksum(schedule[(s_0, j)]
                 for s_0 in range(s)) for j in C for c in C for s in S)

# First semester
model.addConstrs(schedule[(0, c)] * quicksum(pre[c].count(j)
                 for j in C) == 0 for c in C)

# Objective function
model.setObjective(bound, sense=GRB.MINIMIZE)
model.setParam("OutputFlag", 0)

model.update()
model.optimize()

hour_count = 0

# Printing outputs
if model.status == GRB.OPTIMAL:
    print("\nMaximum number of hours required:", model.objVal)
    print("----Schedule---")
    semester = 1
    sem = semester
    sem_hours = 0
    print(f"\nSemester {semester}")
    for v in model.getVars():
        if v.x == 1:
            course = v.VarName[11:20].split("]")[0]
            sem = str(int(v.VarName[9]) + 1)
            print(f"{course}: {hrs[course]}")
            sem_hours += hrs[course]
        if int(sem) > int(semester):
            semester = sem
            print(f"-----Total Hours: {sem_hours}")
            print(f"\nSemester {semester}")
            sem_hours = 0
    # Quick fix for outputting last semester hours
    print(f"-----Total Hours: {sem_hours}")

else:
    print("Not possible :'(")
